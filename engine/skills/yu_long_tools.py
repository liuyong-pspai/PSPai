"""
玉龙自定义工具集 — YuLong Hermes Tools

注册到 Hermes 工具系统，让Agent拥有：
  git_operations  sql_query  csv_read  excel_read  pdf_extract
  self_heal  self_check_gaps  tool_doctor  decision_why  pattern_match
  search_log  workflow_decompose  ssh_exec  brother_watch

依赖：pip install openpyxl pymupdf paramiko pandas
"""

import json
import logging
import os
import re
import sqlite3
import subprocess
import time
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# 工具1：git_operations
# ──────────────────────────────────────────────
def git_operations(action: str, repo_path: str = ".", message: str = "", branch: str = "", remote: str = "origin") -> str:
    """执行Git操作：status/add/commit/push/pull/branch/checkout/log/diff"""
    try:
        if action == "status":
            r = subprocess.run(["git", "status", "--short"], capture_output=True, text=True, cwd=repo_path, timeout=30)
            return r.stdout or r.stderr
        elif action == "add":
            r = subprocess.run(["git", "add", "."], capture_output=True, text=True, cwd=repo_path, timeout=30)
            return r.stdout or "已暂存所有变更"
        elif action == "commit":
            if not message:
                return "❌ commit需要message参数"
            r = subprocess.run(["git", "commit", "-m", message], capture_output=True, text=True, cwd=repo_path, timeout=30)
            return r.stdout or r.stderr
        elif action == "push":
            r = subprocess.run(["git", "push", remote, branch or "HEAD"], capture_output=True, text=True, cwd=repo_path, timeout=120)
            return r.stdout or r.stderr
        elif action == "pull":
            r = subprocess.run(["git", "pull", remote, branch] if branch else ["git", "pull"], capture_output=True, text=True, cwd=repo_path, timeout=120)
            return r.stdout or r.stderr
        elif action == "branch":
            r = subprocess.run(["git", "branch", "-a"], capture_output=True, text=True, cwd=repo_path, timeout=30)
            return r.stdout or r.stderr
        elif action == "checkout":
            if not branch:
                return "❌ checkout需要branch参数"
            r = subprocess.run(["git", "checkout", branch], capture_output=True, text=True, cwd=repo_path, timeout=30)
            return r.stdout or r.stderr
        elif action == "log":
            r = subprocess.run(["git", "log", "--oneline", "-10"], capture_output=True, text=True, cwd=repo_path, timeout=30)
            return r.stdout or r.stderr
        elif action == "diff":
            r = subprocess.run(["git", "diff", "--stat"], capture_output=True, text=True, cwd=repo_path, timeout=30)
            return r.stdout or "无变更"
        else:
            return f"❌ 不支持的操作: {action}。支持: status/add/commit/push/pull/branch/checkout/log/diff"
    except subprocess.TimeoutExpired:
        return "❌ git操作超时"
    except FileNotFoundError:
        return "❌ git未安装或不是git仓库"
    except Exception as e:
        return f"❌ git错误: {e}"


# ──────────────────────────────────────────────
# 工具2：sql_query
# ──────────────────────────────────────────────
def sql_query(query: str, db_path: str = "") -> str:
    """执行SQL查询（自动检测数据库类型：文件存在则为SQLite，以mysql://或postgresql://开头的连接对应类型）"""
    try:
        if db_path.startswith("mysql://") or db_path.startswith("postgresql://"):
            return "⚠️ MySQL/PG支持需要安装pymysql/psycopg2，当前仅支持SQLite"
        if not query:
            return "❌ 需要SQL查询语句"
        db_file = Path(db_path or str(Path.home() / ".hermes-yulong" / "state.db"))
        if not db_file.exists():
            return f"❌ 数据库文件不存在: {db_file}"
        conn = sqlite3.connect(str(db_file))
        conn.row_factory = sqlite3.Row
        cur = conn.execute(query)
        if query.strip().upper().startswith("SELECT") or query.strip().upper().startswith("PRAGMA"):
            rows = cur.fetchall()
            if not rows:
                return "查询结果为空"
            cols = [d[0] for d in cur.description]
            result = " | ".join(cols) + "\n" + "-" * 60
            rows_displayed = 0
            for row in rows:
                if rows_displayed >= 50:
                    break
                vals = [str(row[c])[:30] if row[c] is not None else "" for c in cols]
                result += "\n" + " | ".join(vals)
                rows_displayed += 1
            if len(rows) > 50:
                result += f"\n... 共{len(rows)}行（仅显示前50行）"
            return result
        else:
            conn.commit()
            return f"✅ 影响行数: {cur.rowcount}"
    except sqlite3.Error as e:
        return f"❌ SQL错误: {e}"
    except Exception as e:
        return f"❌ 查询失败: {e}"


# ──────────────────────────────────────────────
# 工具3：csv_read
# ──────────────────────────────────────────────
def csv_read(path: str, max_rows: int = 20, delimiter: str = ",") -> str:
    """读取CSV文件内容并格式化显示"""
    try:
        import csv
        p = Path(path)
        if not p.exists():
            return f"❌ 文件不存在: {path}"
        rows = []
        with open(p, newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for i, row in enumerate(reader):
                if i > max_rows:
                    break
                rows.append(row)
        if not rows:
            return "空文件"
        # 格式化输出
        cols = rows[0]
        result = " | ".join(c[:25] for c in cols) + "\n" + "-" * 60
        for row in rows[1:]:
            result += "\n" + " | ".join((c[:25] if len(c) > 25 else c) for c in row)
        total = len(rows) - 1
        result += f"\n--- 共{total}行（限制显示{max_rows}行）---"
        return result
    except Exception as e:
        return f"❌ CSV读取失败: {e}"


# ──────────────────────────────────────────────
# 工具4：excel_read
# ──────────────────────────────────────────────
def excel_read(path: str, sheet: str = "", max_rows: int = 20) -> str:
    """读取Excel文件内容"""
    try:
        import openpyxl
        p = Path(path)
        if not p.exists():
            return f"❌ 文件不存在: {path}"
        wb = openpyxl.load_workbook(p, data_only=True)
        if sheet:
            ws = wb[sheet]
        else:
            ws = wb.active
            sheet = ws.title
        result = f"[工作表: {sheet}]\n"
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            vals = [str(c or "")[:25] for c in row]
            result += " | ".join(vals) + "\n"
        result += f"--- 共显示{min(max_rows, ws.max_row)}行 ---"
        return result
    except ImportError:
        return "❌ 需要安装 openpyxl: pip install openpyxl"
    except Exception as e:
        return f"❌ Excel读取失败: {e}"


# ──────────────────────────────────────────────
# 工具5：pdf_extract
# ──────────────────────────────────────────────
def pdf_extract(path: str, max_chars: int = 3000) -> str:
    """提取PDF文件文本内容"""
    try:
        import fitz  # PyMuPDF
        p = Path(path)
        if not p.exists():
            return f"❌ 文件不存在: {path}"
        doc = fitz.open(p)
        text = ""
        for page in doc:
            text += page.get_text()
            if len(text) >= max_chars:
                text = text[:max_chars] + "\n... [内容截断]"
                break
        doc.close()
        if not text.strip():
            return "PDF无文本内容（可能是扫描件）"
        return text.strip()
    except ImportError:
        return "❌ 需要安装 PyMuPDF: pip install pymupdf"
    except Exception as e:
        return f"❌ PDF提取失败: {e}"


# ──────────────────────────────────────────────
# 工具6：self_heal
# ──────────────────────────────────────────────
def self_heal() -> str:
    """自检自修：检查自身关键模块完整性并自动修复"""
    results = []
    home = Path(os.environ.get("HERMES_HOME", "~/.hermes-agent"))

    # 1. 检查关键目录
    for d in ["logs", "sessions", "memories", "skills"]:
        p = home / d
        if not p.exists():
            p.mkdir(parents=True, exist_ok=True)
            results.append(f"✅ 重建目录: {d}")
        else:
            results.append(f"✓ {d} 正常")

    # 2. 检查配置文件
    config_file = home / "config.yaml"
    if not config_file.exists():
        results.append("❌ config.yaml 缺失！需要重建")
    else:
        results.append("✓ config.yaml 正常")

    # 3. 检查.env
    env_file = home / ".env"
    if not env_file.exists():
        results.append("❌ .env 缺失！")
    else:
        with open(env_file) as f:
            content = f.read()
        has_feishu = "FEISHU_APP_ID" in content
        has_key = "DEEPSEEK_API_KEY" in content or "API_KEY" in content
        results.append(f"✓ .env 存在（飞书配置{'✓' if has_feishu else '✗'}，API Key{'✓' if has_key else '✗'}）")

    # 4. 检查SOUL.md
    soul = home / "SOUL.md"
    if not soul.exists():
        results.append("⚠️ SOUL.md 缺失")
    else:
        results.append("✓ SOUL.md 正常")

    # 5. 检查磁盘空间
    try:
        st = os.statvfs(str(home))
        free_gb = (st.f_bavail * st.f_frsize) / (1024**3)
        results.append(f"✓ 磁盘剩余: {free_gb:.1f}GB")
    except:
        pass

    return "\n".join(results)


# ──────────────────────────────────────────────
# 工具7：self_check_gaps
# ──────────────────────────────────────────────
def self_check_gaps() -> str:
    """与AI智能体标准能力清单比对，返回缺失项"""
    standard_skills = {
        "文件操作": ["读取", "写入", "搜索", "编辑"],
        "网络": ["网页搜索", "网页抓取", "API调用"],
        "执行": ["Shell命令", "Python代码"],
        "沟通": ["飞书", "发送消息"],
        "记忆": ["读写记忆", "会话搜索"],
        "代码": ["Git操作", "SQL查询", "代码审查"],
        "数据": ["CSV", "Excel", "PDF"],
        "多媒体": ["图片理解", "语音合成", "图片生成"],
        "运维": ["SSH", "自检自修", "进程管理"],
    }
    # 从 TOOLS 注册表动态生成能力清单（不再硬编码）
    _tool_to_category = {
        "git_operations": "代码",
        "sql_query": "代码",
        "csv_read": "数据",
        "excel_read": "数据",
        "pdf_extract": "数据",
        "self_heal": "运维",
        "ssh_exec": "运维",
        "brother_watch": "运维",
    }
    own_skills = {
        "文件操作": ["读取", "写入", "搜索", "编辑"],
        "网络": ["网页搜索", "网页抓取", "API调用"],
        "执行": ["Shell命令", "Python代码"],
        "沟通": ["飞书", "发送消息"],
        "记忆": ["读写记忆", "会话搜索"],
        "代码": [],
        "数据": [],
        "多媒体": ["图片理解"],
        "运维": [],
    }
    for tool_name in TOOLS:
        if tool_name in _tool_to_category:
            cat = _tool_to_category[tool_name]
            desc = TOOLS[tool_name]["description"]
            for skill_item in standard_skills.get(cat, []):
                if skill_item not in own_skills.get(cat, []) and skill_item.lower() in desc.lower():
                    own_skills.setdefault(cat, []).append(skill_item)
    result = "## 能力缺口检查\n\n"
    for category, items in standard_skills.items():
        missing = [s for s in items if s not in own_skills.get(category, [])]
        own = [s for s in items if s in own_skills.get(category, [])]
        if missing:
            result += f"⚠️ **{category}**：缺{'/'.join(missing)}（已有{'/'.join(own) if own else '无'}）\n"
        else:
            result += f"✅ **{category}**：完整\n"
    return result


# ──────────────────────────────────────────────
# 工具8：tool_doctor
# ──────────────────────────────────────────────
def tool_doctor(tool_name: str = "") -> str:
    """检查指定工具或所有工具的健康状态"""
    tools = {
        "git_operations": git_operations,
        "sql_query": sql_query,
        "csv_read": csv_read,
        "excel_read": excel_read,
        "pdf_extract": pdf_extract,
        "self_heal": self_heal,
        "self_check_gaps": self_check_gaps,
        "decision_why": decision_why,
    }
    if tool_name:
        if tool_name not in tools:
            return f"❌ 未知工具: {tool_name}"
        fn = tools[tool_name]
        if not callable(fn):
            return f"❌ {tool_name}: 不是可调用函数"
        import inspect
        try:
            sig = inspect.signature(fn)
            if len(sig.parameters) == 0:
                result = fn()
                return f"✅ {tool_name}: 正常\n{str(result)[:500]}"
            else:
                return f"✅ {tool_name}: 已注册（需{len(sig.parameters)}个参数，未执行调用测试）"
        except Exception as e:
            return f"❌ {tool_name}: {type(e).__name__} - {e}"

    # 检查所有工具
    import inspect
    results = []
    for name, fn in tools.items():
        if not callable(fn):
            results.append(f"❌ {name}: 不是可调用函数")
            continue
        try:
            sig = inspect.signature(fn)
            if len(sig.parameters) == 0:
                r = fn()
                results.append(f"✅ {name}: 正常")
            else:
                results.append(f"✅ {name}: 已注册（需{len(sig.parameters)}个参数）")
        except Exception as e:
            results.append(f"❌ {name}: {type(e).__name__} - {e}")
    return "\n".join(results)


# ──────────────────────────────────────────────
# 工具9：decision_why
# ──────────────────────────────────────────────
def decision_why(question: str) -> str:
    """多维决策分析：收益·风险·替代方案"""
    return (
        f"## 决策分析\n\n"
        f"**问题**：{question}\n\n"
        f"请用标准决策框架分析：\n"
        f"1. **收益分析**：做这件事的正面影响\n"
        f"2. **风险评估**：可能的问题和成本\n"
        f"3. **替代方案**：有没有其他做法\n"
        f"4. **结论建议**：综合判断"
    )


# ──────────────────────────────────────────────
# 工具10：pattern_match
# ──────────────────────────────────────────────
def pattern_match(pattern: str, path: str = ".", file_glob: str = "*.py", max_results: int = 20) -> str:
    """跨文件模式匹配（类似grep）"""
    try:
        import fnmatch
        p = Path(path)
        if not p.exists():
            return f"❌ 路径不存在: {path}"
        results = []
        for f in p.rglob(file_glob):
            if f.is_file() and f.stat().st_size < 1024 * 1024:  # 跳过>1MB
                try:
                    content = f.read_text(encoding="utf-8", errors="replace")
                    for i, line in enumerate(content.split("\n"), 1):
                        if pattern in line and len(results) < max_results:
                            rel = f.relative_to(p)
                            results.append(f"{rel}:{i}: {line.strip()[:100]}")
                except:
                    pass
        if not results:
            return f"未找到匹配 '{pattern}' 在 {path}/{file_glob}"
        return "\n".join(results) + (f"\n... 共{len(results)}条" if len(results) >= max_results else "")
    except Exception as e:
        return f"❌ 模式匹配失败: {e}"


# ──────────────────────────────────────────────
# 工具11：search_log
# ──────────────────────────────────────────────
def search_log(keyword: str, log_dir: str = "~/.hermes-agent/logs", lines: int = 10) -> str:
    """搜索日志文件"""
    try:
        p = Path(log_dir)
        if not p.exists():
            return f"❌ 日志目录不存在: {log_dir}"
        results = []
        for log_file in sorted(p.glob("*.log"), reverse=True):
            try:
                # Stream line-by-line instead of loading entire file
                with open(log_file, "r", encoding="utf-8", errors="replace") as f:
                    for line in f:
                        if keyword in line:
                            results.append(f"[{log_file.name}] {line.strip()[:200]}")
                            if len(results) >= lines:
                                break
            except:
                pass
            if len(results) >= lines:
                break
        if not results:
            return f"日志中未找到: {keyword}"
        return "\n".join(results)
    except Exception as e:
        return f"❌ 日志搜索失败: {e}"


# ──────────────────────────────────────────────
# 工具12：workflow_decompose
# ──────────────────────────────────────────────
def workflow_decompose(goal: str) -> str:
    """将复杂目标分解为可执行步骤"""
    return (
        f"## 工作流分解\n\n"
        f"**目标**：{goal}\n\n"
        f"分解为以下步骤：\n"
        f"1. 分析目标需求\n"
        f"2. 拆分子任务\n"
        f"3. 资源准备\n"
        f"4. 分步执行\n"
        f"5. 验证结果\n"
        f"6. 总结归档\n\n"
        f"请描述具体目标，我将生成详细执行计划。"
    )


# ──────────────────────────────────────────────
# 工具13：ssh_exec (简化版)
# ──────────────────────────────────────────────
HOST_MAP = {
    "M4-1": {
        "host": os.environ.get("M4_1_HOST", "localhost"),
        "user": os.environ.get("M4_1_USER", "yongliu"),
    },
    "M4-2": {
        "host": os.environ.get("M4_2_HOST", "localhost"),
        "user": os.environ.get("M4_2_USER", "yongliu"),
    },
    "DGX": {
        "host": os.environ.get("DGX_HOST", "127.0.0.1"),
        "user": os.environ.get("DGX_USER", "yongliu"),
    },
}

def ssh_exec(host: str, cmd: str) -> str:
    """SSH远程执行命令"""
    host_info = HOST_MAP.get(host.upper(), {})
    if not host_info:
        return f"❌ 未知主机: {host}。可用: {', '.join(HOST_MAP.keys())}"
    if host.upper() == "DGX":
        # 本机执行 — 使用 shell=False 防止命令注入
        try:
            import shlex
            r = subprocess.run(shlex.split(cmd), shell=False, capture_output=True, text=True, timeout=30)
            return (r.stdout or r.stderr)[:2000]
        except subprocess.TimeoutExpired:
            return "❌ 命令超时"
        except Exception as e:
            return f"❌ 执行失败: {e}"
    # 远程
    try:
        import paramiko
        key_path = os.path.expanduser(f"~/.ssh/id_ed25519")
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        privkey = paramiko.Ed25519Key.from_private_key_file(key_path) if os.path.exists(key_path) else None
        if privkey:
            ssh.connect(host_info["host"], username=host_info["user"], pkey=privkey, timeout=10)
        else:
            ssh.connect(host_info["host"], username=host_info["user"], timeout=10)
        _, stdout, stderr = ssh.exec_command(cmd, timeout=30)
        out = stdout.read().decode("utf-8", errors="replace")[:2000]
        err = stderr.read().decode("utf-8", errors="replace")[:500]
        ssh.close()
        if out:
            return out
        return err or "无输出"
    except ImportError:
        return "❌ 需要安装 paramiko: pip install paramiko"
    except Exception as e:
        return f"❌ SSH失败: {e}"


# ──────────────────────────────────────────────
# 工具14：brother_watch
# ──────────────────────────────────────────────
def brother_watch() -> str:
    """监控所有兄弟机器的在线状态"""
    result = "## 兄弟监控状态\n\n"
    for name, info in HOST_MAP.items():
        try:
            r = subprocess.run(
                ["ping", "-c", "1", "-W", "2", info["host"]],
                capture_output=True, text=True, timeout=5
            )
            if r.returncode == 0:
                result += f"✅ {name} ({info['host']}) 在线\n"
            else:
                result += f"❌ {name} ({info['host']}) 离线\n"
        except:
            result += f"⚠️ {name} ({info['host']}) 检测超时\n"
    return result


# ══════════════════════════════════════════════
# Hermes工具注册
# ══════════════════════════════════════════════

TOOLS = {
    "git_operations": {
        "name": "git_operations",
        "description": "执行Git操作：status(状态)/add(暂存)/commit(提交)/push(推送)/pull(拉取)/log(日志)/diff(差异)",
        "parameters": {
            "type": "object",
            "properties": {
                "action": {"type": "string", "description": "操作类型: status/add/commit/push/pull/log/diff"},
                "repo_path": {"type": "string", "description": "Git仓库路径", "default": "."},
                "message": {"type": "string", "description": "commit消息（仅commit操作需要）", "default": ""},
                "branch": {"type": "string", "description": "分支名（仅push/pull操作需要）", "default": ""},
            },
            "required": ["action"],
        },
        "handler": lambda args: git_operations(**args),
    },
    "sql_query": {
        "name": "sql_query",
        "description": "执行SQL查询（支持SQLite）",
        "parameters": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "SQL查询语句"},
                "db_path": {"type": "string", "description": "数据库文件路径", "default": ""},
            },
            "required": ["query"],
        },
        "handler": lambda args: sql_query(**args),
    },
    "csv_read": {
        "name": "csv_read",
        "description": "读取CSV文件内容并格式化显示",
        "parameters": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "CSV文件路径"},
                "max_rows": {"type": "integer", "description": "最大显示行数", "default": 20},
            },
            "required": ["path"],
        },
        "handler": lambda args: csv_read(**args),
    },
    "excel_read": {
        "name": "excel_read",
        "description": "读取Excel文件内容",
        "parameters": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Excel文件路径"},
                "sheet": {"type": "string", "description": "工作表名（默认第一个）", "default": ""},
                "max_rows": {"type": "integer", "description": "最大显示行数", "default": 20},
            },
            "required": ["path"],
        },
        "handler": lambda args: excel_read(**args),
    },
    "pdf_extract": {
        "name": "pdf_extract",
        "description": "提取PDF文件文本内容",
        "parameters": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "PDF文件路径"},
                "max_chars": {"type": "integer", "description": "最大返回字符数", "default": 3000},
            },
            "required": ["path"],
        },
        "handler": lambda args: pdf_extract(**args),
    },
    "self_heal": {
        "name": "self_heal",
        "description": "自检自修：检查自身关键模块完整性并自动修复（目录/配置/密钥/磁盘），无需参数",
        "parameters": {
            "type": "object",
            "properties": {},
        },
        "handler": lambda args: self_heal(),
    },
    "self_check_gaps": {
        "name": "self_check_gaps",
        "description": "与AI智能体标准能力清单比对，返回能力缺口，无需参数",
        "parameters": {
            "type": "object",
            "properties": {},
        },
        "handler": lambda args: self_check_gaps(),
    },
    "tool_doctor": {
        "name": "tool_doctor",
        "description": "检查工具健康状态：指定工具名或检查全部",
        "parameters": {
            "type": "object",
            "properties": {
                "tool_name": {"type": "string", "description": "工具名（空则检查全部）", "default": ""},
            },
        },
        "handler": lambda args: tool_doctor(**args),
    },
    "decision_why": {
        "name": "decision_why",
        "description": "多维决策分析框架：收益/风险/替代方案",
        "parameters": {
            "type": "object",
            "properties": {
                "question": {"type": "string", "description": "决策问题描述"},
            },
            "required": ["question"],
        },
        "handler": lambda args: decision_why(**args),
    },
    "pattern_match": {
        "name": "pattern_match",
        "description": "跨文件模式匹配（类似grep搜索文件内容）",
        "parameters": {
            "type": "object",
            "properties": {
                "pattern": {"type": "string", "description": "搜索关键词"},
                "path": {"type": "string", "description": "搜索路径", "default": "."},
                "file_glob": {"type": "string", "description": "文件通配符（如*.py）", "default": "*.py"},
                "max_results": {"type": "integer", "description": "最大结果数", "default": 20},
            },
            "required": ["pattern"],
        },
        "handler": lambda args: pattern_match(**args),
    },
    "search_log": {
        "name": "search_log",
        "description": "搜索日志文件中包含关键词的行",
        "parameters": {
            "type": "object",
            "properties": {
                "keyword": {"type": "string", "description": "搜索关键词"},
                "log_dir": {"type": "string", "description": "日志目录", "default": "~/.hermes-agent/logs"},
                "lines": {"type": "integer", "description": "返回行数", "default": 10},
            },
            "required": ["keyword"],
        },
        "handler": lambda args: search_log(**args),
    },
    "workflow_decompose": {
        "name": "workflow_decompose",
        "description": "将复杂目标分解为可执行步骤和流程图",
        "parameters": {
            "type": "object",
            "properties": {
                "goal": {"type": "string", "description": "目标描述"},
            },
            "required": ["goal"],
        },
        "handler": lambda args: workflow_decompose(**args),
    },
    "ssh_exec": {
        "name": "ssh_exec",
        "description": "SSH远程执行命令：host=M4-1/M4-2/DGX, cmd=要执行的命令",
        "parameters": {
            "type": "object",
            "properties": {
                "host": {"type": "string", "description": "目标机器: M4-1/M4-2/DGX"},
                "cmd": {"type": "string", "description": "要执行的命令"},
            },
            "required": ["host", "cmd"],
        },
        "handler": lambda args: ssh_exec(**args),
    },
    "brother_watch": {
        "name": "brother_watch",
        "description": "监控所有兄弟机器的在线状态（ping检测），无需参数",
        "parameters": {
            "type": "object",
            "properties": {},
        },
        "handler": lambda args: brother_watch(),
    },
}


def register_all(register_fn) -> None:
    """注册所有工具到Hermes工具系统"""
    for name, tool in TOOLS.items():
        register_fn(
            name=name,
            handler=tool["handler"],
            description=tool["description"],
            parameters=tool["parameters"],
        )
    logger.info(f"✅ 玉龙自定义工具集已注册: {len(TOOLS)}个工具")
